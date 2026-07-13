#!/usr/bin/env python3
"""Local steel basis seasonality dashboard.

This app intentionally uses only the Python standard library plus openpyxl, so
it can run in a light local environment without Streamlit/Plotly.
"""

from __future__ import annotations

import datetime as dt
import json
import os
import re
import shutil
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = Path("/Users/liuziqi/Desktop/daily dashboard/历史数据1.xlsx")
EXCEL_PATH = Path(os.environ.get("STEEL_SPOT_XLSX", str(DEFAULT_EXCEL)))
DB_PATH = Path(os.environ.get("STEEL_BASIS_DB", str(ROOT / "data" / "steel_basis.sqlite")))
DEFAULT_HOST = os.environ.get("STEEL_DASHBOARD_HOST", "0.0.0.0")
DEFAULT_PORT = int(os.environ.get("STEEL_DASHBOARD_PORT", "8765"))
TARGET_MONTHS = (1, 5, 10)
PRODUCT_NAMES = {"HC": "热轧卷板", "RB": "螺纹钢"}
SPOT_SERIES = {
    "shanghai_hotcoil": {
        "label": "上海热卷",
        "excel_column": "上海热卷价格",
        "product": "HC",
    },
    "lecong_hotcoil": {
        "label": "乐从热卷",
        "excel_column": "乐从热卷价格",
        "product": "HC",
    },
    "tangshan_hotcoil": {
        "label": "唐山热卷",
        "excel_column": "唐山热卷价格",
        "product": "HC",
    },
    "zhongtian_factory": {
        "label": "中天厂发",
        "excel_column": "中天厂发价格",
        "product": "RB",
    },
    "center_warehouse": {
        "label": "中心库",
        "excel_column": "中心库价格",
        "product": "RB",
    },
    "guangzhou_shaogang": {
        "label": "广州韶钢",
        "excel_column": "广州韶钢价格",
        "product": "RB",
    },
    "billet": {
        "label": "普方坯",
        "excel_column": "普方坯价格",
        "product": "RB",
    },
    "hangzhou_zhongtian": {
        "label": "杭州中天",
        "excel_column": "杭州中天价格",
        "product": "RB",
    },
}
SPOT_SHEET_NAME = "现货价格"
FUTURES_SHEET_NAME = "期货收盘"
SPOT_HEADERS = ["日期", *[config["excel_column"] for config in SPOT_SERIES.values()]]
FUTURES_HEADERS = ["日期", "品种", "合约月", "合约代码", "收盘价", "来源", "更新时间"]


def now_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")


def connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS spot_prices (
                date TEXT PRIMARY KEY,
                shanghai_hotcoil REAL NOT NULL,
                source TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS spot_series_prices (
                date TEXT NOT NULL,
                spot_key TEXT NOT NULL,
                price REAL NOT NULL,
                source TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (date, spot_key)
            );

            CREATE TABLE IF NOT EXISTS futures_prices (
                date TEXT NOT NULL,
                product TEXT NOT NULL,
                contract_month INTEGER NOT NULL,
                contract_code TEXT NOT NULL,
                close REAL NOT NULL,
                source TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (date, product, contract_code)
            );

            CREATE TABLE IF NOT EXISTS shfe_daily_cache (
                date TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                payload TEXT,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sina_contract_cache (
                contract_code TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                payload TEXT,
                updated_at TEXT NOT NULL
            );
            """
        )


def parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() in {"#N/A", "N/A", "NA", "NULL", "NONE", "-"}:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def excel_error(message: str, path: Path = EXCEL_PATH) -> RuntimeError:
    return RuntimeError(f"{message}：{path}")


def backup_excel(path: Path = EXCEL_PATH) -> str | None:
    if not path.exists():
        return None
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = path.with_suffix(path.suffix + f".{timestamp}.bak")
    try:
        shutil.copy2(path, backup_path)
    except PermissionError as exc:
        raise excel_error("Excel 文件可能正在打开或被占用，备份失败，请关闭后重试", path) from exc
    except OSError as exc:
        raise excel_error(f"Excel 备份失败（{exc}）", path) from exc
    return str(backup_path)


def load_business_workbook(path: Path = EXCEL_PATH) -> Workbook:
    if path.exists():
        try:
            return load_workbook(path)
        except PermissionError as exc:
            raise excel_error("Excel 文件可能正在打开或被占用，读取失败，请关闭后重试", path) from exc
    workbook = Workbook()
    workbook.active.title = SPOT_SHEET_NAME
    return workbook


def row_headers(sheet: Any) -> list[str]:
    return [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]


def ensure_headers(sheet: Any, required_headers: list[str]) -> dict[str, int]:
    headers = row_headers(sheet)
    if not any(headers):
        for idx, header in enumerate(required_headers, start=1):
            sheet.cell(row=1, column=idx, value=header)
        headers = list(required_headers)
    else:
        for header in required_headers:
            if header not in headers:
                sheet.cell(row=1, column=len(headers) + 1, value=header)
                headers.append(header)
    return {header: headers.index(header) + 1 for header in required_headers}


def ensure_excel_structure(workbook: Workbook) -> tuple[Any, Any]:
    if SPOT_SHEET_NAME in workbook.sheetnames:
        spot_sheet = workbook[SPOT_SHEET_NAME]
    else:
        candidate = workbook[workbook.sheetnames[0]]
        candidate_headers = row_headers(candidate)
        if "日期" in candidate_headers or not any(candidate_headers):
            candidate.title = SPOT_SHEET_NAME
            spot_sheet = candidate
        else:
            spot_sheet = workbook.create_sheet(SPOT_SHEET_NAME, 0)

    futures_sheet = (
        workbook[FUTURES_SHEET_NAME]
        if FUTURES_SHEET_NAME in workbook.sheetnames
        else workbook.create_sheet(FUTURES_SHEET_NAME)
    )
    ensure_headers(spot_sheet, SPOT_HEADERS)
    ensure_headers(futures_sheet, FUTURES_HEADERS)
    return spot_sheet, futures_sheet


def save_business_workbook(workbook: Workbook, path: Path = EXCEL_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(path)
    except PermissionError as exc:
        raise excel_error("Excel 文件可能正在打开或被占用，写入失败，请关闭后重试", path) from exc
    except OSError as exc:
        raise excel_error(f"Excel 写入失败（{exc}）", path) from exc


def find_date_row(sheet: Any, date_text: str, date_col: int) -> int | None:
    for row_idx in range(2, sheet.max_row + 1):
        if parse_date(sheet.cell(row=row_idx, column=date_col).value) == date_text:
            return row_idx
    return None


def excel_cell_has_value(sheet: Any, row_idx: int, col_idx: int) -> bool:
    return parse_number(sheet.cell(row=row_idx, column=col_idx).value) is not None


def write_spots_to_excel(
    date_text: str,
    prices_by_spot: dict[str, float],
    path: Path = EXCEL_PATH,
) -> dict[str, Any]:
    workbook = load_business_workbook(path)
    spot_sheet, _futures_sheet = ensure_excel_structure(workbook)
    headers = ensure_headers(spot_sheet, SPOT_HEADERS)
    date_col = headers["日期"]
    row_idx = find_date_row(spot_sheet, date_text, date_col)
    if row_idx is None:
        row_idx = spot_sheet.max_row + 1
        spot_sheet.cell(row=row_idx, column=date_col, value=date_text)

    saved: dict[str, float] = {}
    for spot_key, raw_price in prices_by_spot.items():
        spot_key = normalize_spot_key(spot_key)
        price = parse_number(raw_price)
        if price is None:
            raise_value(f"Invalid spot price for {spot_key}")
        column_name = str(SPOT_SERIES[spot_key]["excel_column"])
        spot_sheet.cell(row=row_idx, column=headers[column_name], value=float(price))
        saved[spot_key] = float(price)

    backup_path = backup_excel(path)
    save_business_workbook(workbook, path)
    return {"saved": saved, "backup": backup_path}


def normalize_futures_product(value: Any, contract_code: str | None = None) -> str:
    text = str(value or "").strip().upper()
    aliases = {"热轧卷板": "HC", "热卷": "HC", "螺纹钢": "RB", "螺纹": "RB"}
    if text in aliases:
        return aliases[text]
    if text in PRODUCT_NAMES:
        return text
    if contract_code:
        prefix = re.match(r"^[A-Z]+", contract_code.upper())
        if prefix and prefix.group(0) in PRODUCT_NAMES:
            return prefix.group(0)
    return normalize_product(text or None)


def month_from_contract(contract_code: str, fallback_month: Any = None) -> int:
    month = parse_number(fallback_month)
    if month is not None:
        return int(month)
    match = re.search(r"(\d{2})$", contract_code.upper())
    if match:
        return int(match.group(1))
    raise_value("Invalid contract month")


def write_futures_to_excel(
    date_text: str,
    product: str,
    contract_month: int,
    contract_code: str,
    close: float,
    source: str,
    path: Path = EXCEL_PATH,
) -> dict[str, Any]:
    product = normalize_product(product)
    contract = contract_code.upper()
    workbook = load_business_workbook(path)
    _spot_sheet, futures_sheet = ensure_excel_structure(workbook)
    headers = ensure_headers(futures_sheet, FUTURES_HEADERS)
    row_idx = None
    for candidate in range(2, futures_sheet.max_row + 1):
        same_date = parse_date(futures_sheet.cell(candidate, headers["日期"]).value) == date_text
        same_product = normalize_futures_product(
            futures_sheet.cell(candidate, headers["品种"]).value,
            str(futures_sheet.cell(candidate, headers["合约代码"]).value or ""),
        ) == product
        same_contract = str(futures_sheet.cell(candidate, headers["合约代码"]).value or "").strip().upper() == contract
        if same_date and same_product and same_contract:
            row_idx = candidate
            break
    if row_idx is None:
        row_idx = futures_sheet.max_row + 1

    values = {
        "日期": date_text,
        "品种": product,
        "合约月": int(contract_month),
        "合约代码": contract,
        "收盘价": float(close),
        "来源": source,
        "更新时间": now_iso(),
    }
    for header, value in values.items():
        futures_sheet.cell(row=row_idx, column=headers[header], value=value)

    backup_path = backup_excel(path)
    save_business_workbook(workbook, path)
    return {"backup": backup_path}


def import_spot_excel(path: Path = EXCEL_PATH) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[SPOT_SHEET_NAME] if SPOT_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]

    try:
        date_idx = headers.index("日期")
    except ValueError as exc:
        raise ValueError("Excel must contain column: 日期") from exc

    series_indexes: dict[str, int] = {}
    missing_columns: list[str] = []
    for spot_key, config in SPOT_SERIES.items():
        column_name = config["excel_column"]
        try:
            series_indexes[spot_key] = headers.index(column_name)
        except ValueError:
            missing_columns.append(column_name)
    if missing_columns:
        raise ValueError(f"Excel missing columns: {', '.join(missing_columns)}")

    imported = 0
    imported_by_series = {spot_key: 0 for spot_key in SPOT_SERIES}
    skipped = 0
    latest_date = None
    with connect() as conn:
        conn.execute("DELETE FROM spot_series_prices")
        conn.execute("DELETE FROM spot_prices")
        for row in rows:
            date_text = parse_date(row[date_idx] if date_idx < len(row) else None)
            if not date_text:
                skipped += 1
                continue

            row_imported = 0
            for spot_key, price_idx in series_indexes.items():
                price = parse_number(row[price_idx] if price_idx < len(row) else None)
                if price is None:
                    continue
                conn.execute(
                    """
                    INSERT INTO spot_series_prices(date, spot_key, price, source, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(date, spot_key) DO UPDATE SET
                        price = excluded.price,
                        source = excluded.source,
                        updated_at = excluded.updated_at
                    """,
                    (date_text, spot_key, price, "excel", now_iso()),
                )
                imported_by_series[spot_key] += 1
                row_imported += 1

                if spot_key == "shanghai_hotcoil":
                    conn.execute(
                        """
                        INSERT INTO spot_prices(date, shanghai_hotcoil, source, updated_at)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(date) DO UPDATE SET
                            shanghai_hotcoil = excluded.shanghai_hotcoil,
                            source = excluded.source,
                            updated_at = excluded.updated_at
                        """,
                        (date_text, price, "excel", now_iso()),
                    )

            if row_imported:
                imported += row_imported
                latest_date = max(latest_date or date_text, date_text)
            else:
                skipped += 1

    return {
        "imported": imported,
        "imported_by_series": imported_by_series,
        "skipped": skipped,
        "latest_date": latest_date,
    }


def import_futures_excel(path: Path = EXCEL_PATH) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if FUTURES_SHEET_NAME not in workbook.sheetnames:
        return {"imported": 0, "skipped": 0, "latest_date": None}

    sheet = workbook[FUTURES_SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    except StopIteration:
        return {"imported": 0, "skipped": 0, "latest_date": None}

    missing = [header for header in FUTURES_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"期货收盘 sheet missing columns: {', '.join(missing)}")
    indexes = {header: headers.index(header) for header in FUTURES_HEADERS}

    imported = 0
    skipped = 0
    latest_date = None
    with connect() as conn:
        conn.execute("DELETE FROM futures_prices")
        for row in rows:
            date_text = parse_date(row[indexes["日期"]] if indexes["日期"] < len(row) else None)
            contract = str(row[indexes["合约代码"]] if indexes["合约代码"] < len(row) else "").strip().upper()
            close = parse_number(row[indexes["收盘价"]] if indexes["收盘价"] < len(row) else None)
            if not date_text or not contract or close is None:
                skipped += 1
                continue
            product = normalize_futures_product(
                row[indexes["品种"]] if indexes["品种"] < len(row) else None,
                contract,
            )
            month = month_from_contract(
                contract,
                row[indexes["合约月"]] if indexes["合约月"] < len(row) else None,
            )
            source = str(row[indexes["来源"]] if indexes["来源"] < len(row) and row[indexes["来源"]] else "excel")
            updated_at = str(
                row[indexes["更新时间"]]
                if indexes["更新时间"] < len(row) and row[indexes["更新时间"]]
                else now_iso()
            )
            save_futures_sqlite(
                date_text,
                product,
                month,
                contract,
                close,
                source,
                updated_at=updated_at,
                conn=conn,
            )
            imported += 1
            latest_date = max(latest_date or date_text, date_text)

    return {"imported": imported, "skipped": skipped, "latest_date": latest_date}


def sync_from_excel(path: Path = EXCEL_PATH) -> dict[str, Any]:
    spot = import_spot_excel(path)
    futures = import_futures_excel(path)
    return {
        "spot": spot,
        "futures": futures,
        "imported": spot["imported"] + futures["imported"],
        "skipped": spot["skipped"] + futures["skipped"],
    }


def sync_sqlite_business_data_to_excel(path: Path = EXCEL_PATH) -> dict[str, Any]:
    """One-time migration: merge current SQLite business rows back into Excel."""
    init_db()
    workbook = load_business_workbook(path)
    spot_sheet, futures_sheet = ensure_excel_structure(workbook)
    spot_headers = ensure_headers(spot_sheet, SPOT_HEADERS)
    futures_headers = ensure_headers(futures_sheet, FUTURES_HEADERS)

    spot_rows_by_date: dict[str, int] = {}
    for row_idx in range(2, spot_sheet.max_row + 1):
        date_text = parse_date(spot_sheet.cell(row=row_idx, column=spot_headers["日期"]).value)
        if date_text:
            spot_rows_by_date[date_text] = row_idx

    futures_rows_by_key: dict[tuple[str, str, str], int] = {}
    for row_idx in range(2, futures_sheet.max_row + 1):
        date_text = parse_date(futures_sheet.cell(row=row_idx, column=futures_headers["日期"]).value)
        contract = str(futures_sheet.cell(row=row_idx, column=futures_headers["合约代码"]).value or "").strip().upper()
        if not date_text or not contract:
            continue
        product = normalize_futures_product(
            futures_sheet.cell(row=row_idx, column=futures_headers["品种"]).value,
            contract,
        )
        futures_rows_by_key[(date_text, product, contract)] = row_idx

    spot_written = 0
    futures_written = 0
    with connect() as conn:
        spot_rows = conn.execute(
            """
            SELECT date, spot_key, price, source
            FROM spot_series_prices
            ORDER BY date, spot_key
            """
        ).fetchall()
        futures_rows = conn.execute(
            """
            SELECT date, product, contract_month, contract_code, close, source, updated_at
            FROM futures_prices
            ORDER BY date, product, contract_code
            """
        ).fetchall()

    for row in spot_rows:
        spot_key = normalize_spot_key(row["spot_key"])
        date_text = parse_date(row["date"])
        if not date_text:
            continue
        row_idx = spot_rows_by_date.get(date_text)
        if row_idx is None:
            row_idx = spot_sheet.max_row + 1
            spot_sheet.cell(row=row_idx, column=spot_headers["日期"], value=date_text)
            spot_rows_by_date[date_text] = row_idx
        col_idx = spot_headers[str(SPOT_SERIES[spot_key]["excel_column"])]
        should_write = row["source"] == "manual" or not excel_cell_has_value(spot_sheet, row_idx, col_idx)
        if should_write:
            spot_sheet.cell(row=row_idx, column=col_idx, value=float(row["price"]))
            spot_written += 1

    for row in futures_rows:
        date_text = parse_date(row["date"])
        if not date_text:
            continue
        product = normalize_product(row["product"])
        contract = str(row["contract_code"]).upper()
        key = (date_text, product, contract)
        row_idx = futures_rows_by_key.get(key)
        if row_idx is None:
            row_idx = futures_sheet.max_row + 1
            futures_rows_by_key[key] = row_idx
        values = {
            "日期": date_text,
            "品种": product,
            "合约月": int(row["contract_month"]),
            "合约代码": contract,
            "收盘价": float(row["close"]),
            "来源": row["source"] or "sqlite",
            "更新时间": row["updated_at"] or now_iso(),
        }
        for header, value in values.items():
            futures_sheet.cell(row=row_idx, column=futures_headers[header], value=value)
        futures_written += 1

    backup_path = backup_excel(path)
    save_business_workbook(workbook, path)
    sync_result = sync_from_excel(path)
    return {
        "backup": backup_path,
        "spot_written": spot_written,
        "futures_written": futures_written,
        "sync": sync_result,
    }


def next_contract_code(date_text: str, product: str, month: int) -> str:
    trade_date = dt.date.fromisoformat(date_text)
    year = trade_date.year if trade_date.month <= month else trade_date.year + 1
    return f"{product.upper()}{str(year)[-2:]}{month:02d}"


def contract_window(contract_code: str, fallback_month: int) -> tuple[dt.date, dt.date, str]:
    match = re.search(r"(\d{2})(\d{2})$", contract_code.upper())
    if match:
        contract_year = 2000 + int(match.group(1))
        month = int(match.group(2))
    else:
        contract_year = dt.date.today().year
        month = int(fallback_month)
    start = dt.date(contract_year - 1, month, 15)
    end = dt.date(contract_year, month, 15)
    return start, end, f"{start.year}-{end.year}"


def normalize_product(value: str | None) -> str:
    product = (value or "HC").upper()
    if product not in PRODUCT_NAMES:
        raise ValueError("product must be HC or RB")
    return product


def normalize_spot_key(value: str | None) -> str:
    spot_key = value or "shanghai_hotcoil"
    if spot_key not in SPOT_SERIES:
        raise ValueError("Unknown spot series")
    return spot_key


def product_for_spot(spot_key: str) -> str:
    return str(SPOT_SERIES[normalize_spot_key(spot_key)]["product"])


def shfe_daily_url(date_text: str) -> str:
    return (
        "https://www.shfe.com.cn/data/tradedata/future/dailydata/"
        f"kx{date_text.replace('-', '')}.dat"
    )


def sina_daily_url(contract_code: str) -> str:
    return (
        "https://stock.finance.sina.com.cn/futures/api/json.php/"
        "IndexService.getInnerFuturesDailyKLine?"
        f"symbol={urllib.parse.quote(contract_code.upper())}"
    )


def fetch_sina_contract_daily(contract_code: str, use_cache: bool = True) -> list[dict[str, Any]]:
    contract = contract_code.upper()
    with connect() as conn:
        if use_cache:
            row = conn.execute(
                "SELECT status, payload FROM sina_contract_cache WHERE contract_code = ?",
                (contract,),
            ).fetchone()
            if row:
                if row["status"] == "ok":
                    return json.loads(row["payload"] or "[]")
                if row["status"] == "empty":
                    raise RuntimeError(f"Sina Futures has no cached daily rows for {contract}")

        request = urllib.request.Request(
            sina_daily_url(contract),
            headers={
                "User-Agent": "Mozilla/5.0 steel-basis-dashboard/1.0",
                "Accept": "application/json,text/plain,*/*",
                "Referer": "https://finance.sina.com.cn/futures/",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=12) as response:
                raw = response.read().decode("utf-8", errors="replace").strip()
        except urllib.error.HTTPError as exc:
            conn.execute(
                "INSERT OR REPLACE INTO sina_contract_cache(contract_code, status, payload, updated_at) VALUES (?, ?, ?, ?)",
                (contract, f"http_{exc.code}", None, now_iso()),
            )
            raise RuntimeError(f"Sina HTTP {exc.code} for {contract}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"Cannot reach Sina Futures: {exc.reason}") from exc

        if raw.startswith("/*"):
            raw = raw.split("*/", 1)[-1].strip()
        if raw.startswith("var "):
            raw = raw.split("=", 1)[-1].strip().rstrip(";")

        payload = json.loads(raw)
        rows: list[dict[str, Any]] = []
        for item in payload or []:
            if isinstance(item, list) and len(item) >= 5:
                date_text = parse_date(item[0])
                close = parse_number(item[4])
                if date_text and close is not None:
                    rows.append({"date": date_text, "close": close})
            elif isinstance(item, dict):
                date_text = parse_date(item.get("d") or item.get("date"))
                close = parse_number(item.get("c") or item.get("close"))
                if date_text and close is not None:
                    rows.append({"date": date_text, "close": close})

        status = "ok" if rows else "empty"
        conn.execute(
            "INSERT OR REPLACE INTO sina_contract_cache(contract_code, status, payload, updated_at) VALUES (?, ?, ?, ?)",
            (contract, status, json.dumps(rows, ensure_ascii=False), now_iso()),
        )
        if not rows:
            raise RuntimeError(f"Sina Futures returned no daily rows for {contract}")
        return rows


def fetch_sina_contract_close(date_text: str, contract_code: str) -> float:
    rows = fetch_sina_contract_daily(contract_code)
    for row in rows:
        if row["date"] == date_text:
            return float(row["close"])
    raise RuntimeError(f"{contract_code.upper()} has no Sina close price for {date_text}")


def fetch_shfe_daily(date_text: str, use_cache: bool = True) -> list[dict[str, Any]]:
    with connect() as conn:
        if use_cache:
            row = conn.execute(
                "SELECT status, payload FROM shfe_daily_cache WHERE date = ?", (date_text,)
            ).fetchone()
            if row:
                if row["status"] == "ok":
                    return json.loads(row["payload"] or "[]")
                if row["status"] == "empty":
                    raise RuntimeError(f"SHFE has no cached daily data for {date_text}")

        request = urllib.request.Request(
            shfe_daily_url(date_text),
            headers={
                "User-Agent": "Mozilla/5.0 steel-basis-dashboard/1.0",
                "Accept": "application/json,text/plain,*/*",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=12) as response:
                raw = response.read().decode("utf-8", errors="replace")
        except urllib.error.HTTPError as exc:
            conn.execute(
                "INSERT OR REPLACE INTO shfe_daily_cache(date, status, payload, updated_at) VALUES (?, ?, ?, ?)",
                (date_text, f"http_{exc.code}", None, now_iso()),
            )
            raise RuntimeError(f"SHFE HTTP {exc.code} for {date_text}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"Cannot reach SHFE: {exc.reason}") from exc

        payload = json.loads(raw)
        rows = payload.get("o_curinstrument") or payload.get("data") or []
        status = "ok" if rows else "empty"
        conn.execute(
            "INSERT OR REPLACE INTO shfe_daily_cache(date, status, payload, updated_at) VALUES (?, ?, ?, ?)",
            (date_text, status, json.dumps(rows, ensure_ascii=False), now_iso()),
        )
        if not rows:
            raise RuntimeError(f"SHFE returned no daily rows for {date_text}")
        return rows


def row_instrument(row: dict[str, Any]) -> str:
    for key in ("INSTRUMENTID", "instrumentid", "InstrumentID", "合约代码"):
        value = row.get(key)
        if value:
            return str(value).strip().upper()
    product = row.get("PRODUCTGROUPID") or row.get("productgroupid")
    month = row.get("DELIVERYMONTH") or row.get("deliverymonth")
    if product and month:
        return f"{str(product).strip()}{str(month).strip()}".upper()
    for key, value in row.items():
        if "instrument" in str(key).lower() and value:
            return str(value).strip().upper()
    return ""


def row_close(row: dict[str, Any]) -> float | None:
    for key in ("CLOSEPRICE", "closeprice", "ClosePrice", "CLOSE", "close", "收盘价"):
        if key in row:
            return parse_number(row.get(key))
    return None


def close_from_daily_rows(rows: list[dict[str, Any]], contract_code: str) -> float:
    target = contract_code.upper()
    for row in rows:
        if row_instrument(row) == target:
            close = row_close(row)
            if close is None:
                raise RuntimeError(f"SHFE row for {target} has no close price")
            return close
    date_hint = "this date"
    raise RuntimeError(f"{target} was not found in SHFE daily data for {date_hint}")


def fetch_contract_close(date_text: str, contract_code: str) -> float:
    close, _source = fetch_contract_close_with_source(date_text, contract_code)
    return close


def fetch_contract_close_with_source(date_text: str, contract_code: str) -> tuple[float, str]:
    errors: list[str] = []
    try:
        return fetch_sina_contract_close(date_text, contract_code), "sina"
    except Exception as exc:
        errors.append(str(exc))

    try:
        rows = fetch_shfe_daily(date_text)
        return close_from_daily_rows(rows, contract_code), "shfe"
    except RuntimeError as exc:
        if "was not found" in str(exc):
            errors.append(f"{contract_code.upper()} was not found in SHFE daily data for {date_text}")
            raise RuntimeError("；".join(errors)) from exc
        raise
    except Exception as exc:
        errors.append(str(exc))
        raise RuntimeError("；".join(errors)) from exc


def save_spot(
    date_text: str,
    price: float,
    source: str = "manual",
    spot_key: str = "shanghai_hotcoil",
) -> None:
    save_spots(date_text, {spot_key: price}, source=source)


def save_spots(
    date_text: str,
    prices_by_spot: dict[str, float],
    source: str = "manual",
) -> dict[str, Any]:
    date_text = parse_date(date_text) or raise_value("Invalid date")
    cleaned: dict[str, float] = {}
    for spot_key, raw_price in prices_by_spot.items():
        spot_key = normalize_spot_key(spot_key)
        price = parse_number(raw_price)
        if price is None:
            raise_value(f"Invalid spot price for {spot_key}")
        cleaned[spot_key] = float(price)
    if not cleaned:
        raise_value("No spot prices to save")

    excel_result = write_spots_to_excel(date_text, cleaned)
    with connect() as conn:
        for spot_key, price in cleaned.items():
            save_spot_sqlite(date_text, price, source=source, spot_key=spot_key, conn=conn)
    return {"saved": excel_result["saved"], "backup": excel_result["backup"]}


def save_spot_sqlite(
    date_text: str,
    price: float,
    source: str = "manual",
    spot_key: str = "shanghai_hotcoil",
    updated_at: str | None = None,
    conn: sqlite3.Connection | None = None,
) -> None:
    date_text = parse_date(date_text) or raise_value("Invalid date")
    spot_key = normalize_spot_key(spot_key)
    timestamp = updated_at or now_iso()

    def write(target: sqlite3.Connection) -> None:
        target.execute(
            """
            INSERT INTO spot_series_prices(date, spot_key, price, source, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(date, spot_key) DO UPDATE SET
                price = excluded.price,
                source = excluded.source,
                updated_at = excluded.updated_at
            """,
            (date_text, spot_key, float(price), source, timestamp),
        )
        if spot_key == "shanghai_hotcoil":
            target.execute(
                """
                INSERT INTO spot_prices(date, shanghai_hotcoil, source, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(date) DO UPDATE SET
                    shanghai_hotcoil = excluded.shanghai_hotcoil,
                    source = excluded.source,
                    updated_at = excluded.updated_at
                """,
                (date_text, float(price), source, timestamp),
            )

    if conn is None:
        with connect() as owned_conn:
            write(owned_conn)
    else:
        write(conn)


def save_futures(
    date_text: str,
    product: str,
    contract_month: int,
    contract_code: str,
    close: float,
    source: str,
) -> None:
    product = normalize_product(product)
    date_text = parse_date(date_text) or raise_value("Invalid date")
    contract = contract_code.upper()
    write_futures_to_excel(date_text, product, int(contract_month), contract, float(close), source)
    save_futures_sqlite(date_text, product, int(contract_month), contract, float(close), source)


def save_futures_sqlite(
    date_text: str,
    product: str,
    contract_month: int,
    contract_code: str,
    close: float,
    source: str,
    updated_at: str | None = None,
    conn: sqlite3.Connection | None = None,
) -> None:
    product = normalize_product(product)
    date_text = parse_date(date_text) or raise_value("Invalid date")
    timestamp = updated_at or now_iso()

    def write(target: sqlite3.Connection) -> None:
        target.execute(
            """
            INSERT INTO futures_prices(date, product, contract_month, contract_code, close, source, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(date, product, contract_code) DO UPDATE SET
                close = excluded.close,
                source = excluded.source,
                updated_at = excluded.updated_at
            """,
            (
                date_text,
                product,
                int(contract_month),
                contract_code.upper(),
                float(close),
                source,
                timestamp,
            ),
        )

    if conn is None:
        with connect() as owned_conn:
            write(owned_conn)
    else:
        write(conn)


def raise_value(message: str) -> None:
    raise ValueError(message)


def latest_spot_date(spot_key: str = "shanghai_hotcoil") -> str | None:
    spot_key = normalize_spot_key(spot_key)
    with connect() as conn:
        row = conn.execute(
            "SELECT MAX(date) AS date FROM spot_series_prices WHERE spot_key = ?",
            (spot_key,),
        ).fetchone()
    return row["date"] if row and row["date"] else None


def latest_trading_date_from_shfe(start_date: str | None = None) -> str:
    start = dt.date.fromisoformat(start_date) if start_date else dt.date.today()
    for offset in range(0, 16):
        candidate = start - dt.timedelta(days=offset)
        if candidate.weekday() >= 5:
            continue
        date_text = candidate.isoformat()
        try:
            fetch_shfe_daily(date_text)
            return date_text
        except Exception:
            continue
    latest = latest_spot_date()
    if latest:
        return latest
    return start.isoformat()


def state_payload(spot_key: str = "shanghai_hotcoil") -> dict[str, Any]:
    init_db()
    spot_key = normalize_spot_key(spot_key)
    product = product_for_spot(spot_key)
    with connect() as conn:
        spot_count = conn.execute(
            "SELECT COUNT(*) AS n FROM spot_series_prices WHERE spot_key = ?",
            (spot_key,),
        ).fetchone()["n"]
        futures_count = conn.execute(
            "SELECT COUNT(*) AS n FROM futures_prices WHERE product = ?",
            (product,),
        ).fetchone()["n"]
        latest_spot = conn.execute(
            """
            SELECT date, price, source
            FROM spot_series_prices
            WHERE spot_key = ?
            ORDER BY date DESC
            LIMIT 1
            """,
            (spot_key,),
        ).fetchone()
        latest_futures = conn.execute(
            """
            SELECT date, product, contract_month, contract_code, close, source
            FROM futures_prices
            WHERE product = ?
            ORDER BY date DESC, product, contract_month
            LIMIT 12
            """,
            (product,),
        ).fetchall()
        year_rows = conn.execute(
            """
            SELECT DISTINCT substr(date, 1, 4) AS year
            FROM spot_series_prices
            WHERE spot_key = ?
            ORDER BY year
            """,
            (spot_key,),
        ).fetchall()

    today = dt.date.today().isoformat()
    default_date = latest_spot["date"] if latest_spot else today
    return {
        "excel_path": str(EXCEL_PATH),
        "db_path": str(DB_PATH),
        "spot_count": spot_count,
        "futures_count": futures_count,
        "latest_spot": dict(latest_spot) if latest_spot else None,
        "latest_futures": [dict(row) for row in latest_futures],
        "years": [row["year"] for row in year_rows],
        "default_date": default_date,
        "default_contracts": {
            f"{m:02d}": next_contract_code(default_date, product, m) for m in TARGET_MONTHS
        },
        "months": [f"{m:02d}" for m in TARGET_MONTHS],
        "products": PRODUCT_NAMES,
        "spot_key": spot_key,
        "spot_series": SPOT_SERIES,
        "product": product,
    }


def basis_payload(
    product: str | None,
    contract_month: int,
    spot_key: str = "shanghai_hotcoil",
) -> dict[str, Any]:
    spot_key = normalize_spot_key(spot_key)
    product = normalize_product(product or product_for_spot(spot_key))
    month = int(contract_month)
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT
                s.date,
                s.price AS spot,
                f.close AS futures_close,
                f.contract_code,
                f.source AS futures_source,
                s.source AS spot_source
            FROM spot_series_prices s
            JOIN futures_prices f
                ON f.date = s.date
            WHERE f.product = ? AND f.contract_month = ?
                AND s.spot_key = ?
            ORDER BY s.date
            """,
            (product, month, spot_key),
        ).fetchall()

    points = []
    for row in rows:
        date_value = dt.date.fromisoformat(row["date"])
        window_start, window_end, season_label = contract_window(row["contract_code"], month)
        if date_value < window_start or date_value > window_end:
            continue
        points.append(
            {
                "date": row["date"],
                "year": window_start.year,
                "season_label": season_label,
                "month_day": date_value.strftime("%m-%d"),
                "day_index": int(date_value.strftime("%j")),
                "window_day": (date_value - window_start).days,
                "window_total_days": (window_end - window_start).days,
                "window_start": window_start.isoformat(),
                "window_end": window_end.isoformat(),
                "spot": row["spot"],
                "futures_close": row["futures_close"],
                "basis": row["spot"] - row["futures_close"],
                "contract_code": row["contract_code"],
                "spot_source": row["spot_source"],
                "futures_source": row["futures_source"],
            }
        )
    return {
        "product": product,
        "spot_key": spot_key,
        "spot_label": SPOT_SERIES[spot_key]["label"],
        "contract_month": f"{month:02d}",
        "points": points,
    }


def recent_rows(limit: int = 12, spot_key: str = "shanghai_hotcoil") -> list[dict[str, Any]]:
    spot_key = normalize_spot_key(spot_key)
    product = product_for_spot(spot_key)
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT
                s.date,
                s.price,
                GROUP_CONCAT(f.product || f.contract_month || ':' || f.contract_code || '=' || f.close, '; ') AS futures
            FROM spot_series_prices s
            LEFT JOIN futures_prices f
                ON f.date = s.date
                AND f.product = ?
            WHERE s.spot_key = ?
            GROUP BY s.date, s.price
            ORDER BY s.date DESC
            LIMIT ?
            """,
            (product, spot_key, limit),
        ).fetchall()
    return [dict(row) for row in rows]


def recent_overview_rows(limit: int = 5) -> list[dict[str, Any]]:
    spot_keys = tuple(SPOT_SERIES.keys())
    placeholders = ",".join("?" for _ in spot_keys)
    with connect() as conn:
        date_rows = conn.execute(
            f"""
            SELECT date
            FROM spot_series_prices
            WHERE spot_key IN ({placeholders})
            GROUP BY date
            HAVING COUNT(DISTINCT spot_key) = ?
            ORDER BY date DESC
            LIMIT ?
            """,
            (*spot_keys, len(spot_keys), int(limit)),
        ).fetchall()
        dates = [row["date"] for row in date_rows]
        if not dates:
            return []

        date_placeholders = ",".join("?" for _ in dates)
        spot_rows = conn.execute(
            f"""
            SELECT date, spot_key, price
            FROM spot_series_prices
            WHERE date IN ({date_placeholders})
                AND spot_key IN ({placeholders})
            """,
            (*dates, *spot_keys),
        ).fetchall()

        futures_rows = conn.execute(
            f"""
            SELECT date, product, contract_code, close
            FROM futures_prices
            WHERE date IN ({date_placeholders})
                AND product IN ('HC', 'RB')
            """,
            (*dates,),
        ).fetchall()

    spot_by_date = {
        date_text: {spot_key: None for spot_key in spot_keys}
        for date_text in dates
    }
    for row in spot_rows:
        spot_by_date[row["date"]][row["spot_key"]] = row["price"]

    futures_by_key = {
        (row["date"], row["product"], row["contract_code"]): row["close"]
        for row in futures_rows
    }

    result: list[dict[str, Any]] = []
    for date_text in dates:
        row_payload = {"date": date_text, "series": {}}
        for spot_key in spot_keys:
            product = product_for_spot(spot_key)
            price = spot_by_date[date_text].get(spot_key)
            series_payload = {
                "label": SPOT_SERIES[spot_key]["label"],
                "product": product,
                "price": price,
                "basis": {},
            }
            for month in TARGET_MONTHS:
                contract = next_contract_code(date_text, product, month)
                close = futures_by_key.get((date_text, product, contract))
                series_payload["basis"][f"{month:02d}"] = {
                    "contract_code": contract,
                    "close": close,
                    "basis": None if price is None or close is None else price - close,
                }
            row_payload["series"][spot_key] = series_payload
        result.append(row_payload)
    return result


def fetch_missing(
    product: str,
    month: int,
    limit: int,
    spot_key: str = "shanghai_hotcoil",
) -> dict[str, Any]:
    spot_key = normalize_spot_key(spot_key)
    product = normalize_product(product or product_for_spot(spot_key))
    month = int(month)
    attempted = 0
    saved = 0
    error_count = 0
    errors: list[str] = []
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT s.date
            FROM spot_series_prices s
            LEFT JOIN futures_prices f
                ON f.date = s.date
                AND f.product = ?
                AND f.contract_month = ?
            WHERE f.date IS NULL
                AND s.spot_key = ?
            ORDER BY s.date DESC
            LIMIT ?
            """,
            (product, month, spot_key, int(limit)),
        ).fetchall()

    for row in rows:
        date_text = row["date"]
        contract = next_contract_code(date_text, product, month)
        attempted += 1
        try:
            close, source = fetch_contract_close_with_source(date_text, contract)
            save_futures(date_text, product, month, contract, close, source)
            saved += 1
            time.sleep(0.15)
        except Exception as exc:
            error_count += 1
            if len(errors) < 8:
                errors.append(f"{date_text} {contract}: {exc}")
    return {"attempted": attempted, "saved": saved, "error_count": error_count, "errors": errors}


def fetch_missing_all_months(
    product: str,
    limit: int,
    spot_key: str = "shanghai_hotcoil",
) -> dict[str, Any]:
    spot_key = normalize_spot_key(spot_key)
    product = normalize_product(product or product_for_spot(spot_key))
    attempted_dates = 0
    saved = 0
    skipped_existing = 0
    error_count = 0
    errors: list[str] = []
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT s.date
            FROM spot_series_prices s
            WHERE EXISTS (
                SELECT 1
                FROM (SELECT 1 AS m UNION ALL SELECT 5 UNION ALL SELECT 10) months
                LEFT JOIN futures_prices f
                    ON f.date = s.date
                    AND f.product = ?
                    AND f.contract_month = months.m
                WHERE f.date IS NULL
            )
                AND s.spot_key = ?
            ORDER BY s.date DESC
            LIMIT ?
            """,
            (product, spot_key, int(limit)),
        ).fetchall()

    for row in rows:
        date_text = row["date"]
        attempted_dates += 1

        with connect() as conn:
            existing = {
                existing_row["contract_month"]
                for existing_row in conn.execute(
                    "SELECT contract_month FROM futures_prices WHERE date = ? AND product = ?",
                    (date_text, product),
                ).fetchall()
            }

        for month in TARGET_MONTHS:
            if month in existing:
                skipped_existing += 1
                continue
            contract = next_contract_code(date_text, product, month)
            try:
                close, source = fetch_contract_close_with_source(date_text, contract)
                save_futures(date_text, product, month, contract, close, source)
                saved += 1
            except Exception as exc:
                error_count += 1
                if len(errors) < 8:
                    errors.append(f"{date_text} {contract}: {exc}")
        time.sleep(0.15)

    return {
        "attempted_dates": attempted_dates,
        "saved": saved,
        "skipped_existing": skipped_existing,
        "error_count": error_count,
        "errors": errors,
    }


class AppHandler(BaseHTTPRequestHandler):
    server_version = "SteelBasisDashboard/1.0"

    def log_message(self, fmt: str, *args: Any) -> None:
        sys.stderr.write("%s - %s\n" % (self.address_string(), fmt % args))

    def send_json(self, payload: Any, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_html(self) -> None:
        body = HTML.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8") if length else "{}"
        return json.loads(raw or "{}")

    def do_GET(self) -> None:
        try:
            path = self.path.split("?", 1)[0]
            query = parse_query(self.path)
            if path == "/":
                self.send_html()
            elif path == "/api/state":
                self.send_json(state_payload(query.get("spot_key", "shanghai_hotcoil")))
            elif path == "/api/basis":
                self.send_json(
                    basis_payload(
                        query.get("product"),
                        int(query.get("contract_month", "5")),
                        query.get("spot_key", "shanghai_hotcoil"),
                    )
                )
            elif path == "/api/recent":
                self.send_json(
                    {
                        "rows": recent_rows(
                            int(query.get("limit", "12")),
                            query.get("spot_key", "shanghai_hotcoil"),
                        )
                    }
                )
            elif path == "/api/recent-overview":
                self.send_json({"rows": recent_overview_rows(int(query.get("limit", "5")))})
            elif path == "/api/latest-trading-date":
                self.send_json({"date": latest_trading_date_from_shfe(query.get("start"))})
            else:
                self.send_error(HTTPStatus.NOT_FOUND, "Not found")
        except Exception as exc:
            self.send_json({"error": str(exc)}, 500)

    def do_POST(self) -> None:
        try:
            path = self.path.split("?", 1)[0]
            payload = self.read_json()
            if path == "/api/import-excel":
                self.send_json(sync_from_excel())
            elif path == "/api/spot":
                date_text = parse_date(payload.get("date")) or raise_value("Invalid date")
                price = parse_number(payload.get("price"))
                if price is None:
                    raise_value("Invalid spot price")
                save_spot(date_text, price, spot_key=payload.get("spot_key", "shanghai_hotcoil"))
                self.send_json({"ok": True, "date": date_text, "price": price})
            elif path == "/api/spots":
                date_text = parse_date(payload.get("date")) or raise_value("Invalid date")
                prices = payload.get("prices") or {}
                cleaned = {}
                for spot_key, raw_price in prices.items():
                    if raw_price in (None, ""):
                        continue
                    price = parse_number(raw_price)
                    if price is None:
                        raise_value(f"Invalid spot price for {spot_key}")
                    cleaned[spot_key] = price
                if not cleaned:
                    raise_value("No spot prices to save")
                result = save_spots(date_text, cleaned)
                self.send_json({"ok": True, "date": date_text, **result})
            elif path == "/api/futures":
                date_text = parse_date(payload.get("date")) or raise_value("Invalid date")
                product = normalize_product(payload.get("product"))
                month = int(payload.get("contract_month"))
                contract = str(
                    payload.get("contract_code") or next_contract_code(date_text, product, month)
                ).upper()
                close = parse_number(payload.get("close"))
                if close is None:
                    raise_value("Invalid futures close")
                save_futures(date_text, product, month, contract, close, "manual")
                self.send_json({"ok": True, "date": date_text, "contract_code": contract, "close": close})
            elif path == "/api/fetch-futures":
                date_text = parse_date(payload.get("date")) or raise_value("Invalid date")
                product = normalize_product(payload.get("product"))
                month = int(payload.get("contract_month"))
                contract = str(
                    payload.get("contract_code") or next_contract_code(date_text, product, month)
                ).upper()
                close, source = fetch_contract_close_with_source(date_text, contract)
                save_futures(date_text, product, month, contract, close, source)
                self.send_json(
                    {
                        "ok": True,
                        "date": date_text,
                        "contract_code": contract,
                        "close": close,
                        "source": source,
                    }
                )
            elif path == "/api/fetch-missing":
                self.send_json(
                    fetch_missing(
                        payload.get("product", "HC"),
                        int(payload.get("contract_month", 5)),
                        int(payload.get("limit", 20)),
                        payload.get("spot_key", "shanghai_hotcoil"),
                    )
                )
            elif path == "/api/fetch-missing-all":
                self.send_json(
                    fetch_missing_all_months(
                        payload.get("product", "HC"),
                        int(payload.get("limit", 20)),
                        payload.get("spot_key", "shanghai_hotcoil"),
                    )
                )
            else:
                self.send_error(HTTPStatus.NOT_FOUND, "Not found")
        except Exception as exc:
            self.send_json({"error": str(exc)}, 400)


def parse_query(path: str) -> dict[str, str]:
    if "?" not in path:
        return {}
    query = path.split("?", 1)[1]
    result: dict[str, str] = {}
    for part in query.split("&"):
        if not part:
            continue
        key, _, value = part.partition("=")
        result[urllib.parse.unquote_plus(key)] = urllib.parse.unquote_plus(value)
    return result


HTML = r"""
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>上海热卷基差季节性看板</title>
  <style>
    :root {
      color-scheme: light;
      --ink: #202124;
      --muted: #667085;
      --line: #d9dee7;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --blue: #1f6feb;
      --green: #248a3d;
      --red: #c2410c;
      --yellow: #b7791f;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: var(--bg);
      color: var(--ink);
    }
    header {
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      padding: 18px 24px 14px;
    }
    h1 {
      margin: 0;
      font-size: 22px;
      line-height: 1.25;
      letter-spacing: 0;
    }
    main {
      max-width: 1360px;
      margin: 0 auto;
      padding: 18px 20px 28px;
    }
    .topbar {
      display: none;
      grid-template-columns: repeat(4, minmax(150px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }
    .metric, .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .metric {
      padding: 12px 14px;
      min-height: 76px;
    }
    .metric .label {
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 8px;
    }
    .metric .value {
      font-size: 22px;
      font-weight: 700;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    .layout {
      display: grid;
      grid-template-columns: minmax(0, 1fr) 380px;
      gap: 16px;
      align-items: start;
    }
    .panel {
      padding: 14px;
    }
    .panel h2 {
      margin: 0 0 12px;
      font-size: 15px;
      letter-spacing: 0;
    }
    .chart-toolbar, .row {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
    }
    .chart-toolbar {
      justify-content: space-between;
      margin-bottom: 10px;
    }
    label {
      display: block;
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 5px;
    }
    input, select, button {
      height: 34px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      padding: 0 10px;
      font: inherit;
      font-size: 13px;
    }
    button {
      cursor: pointer;
      font-weight: 600;
    }
    button.primary {
      background: var(--blue);
      border-color: var(--blue);
      color: white;
    }
    button:disabled {
      cursor: wait;
      opacity: .65;
    }
    .segmented {
      display: inline-flex;
      border: 1px solid var(--line);
      border-radius: 7px;
      overflow: hidden;
      background: #fff;
    }
    .segmented button {
      border: 0;
      border-right: 1px solid var(--line);
      border-radius: 0;
      min-width: 52px;
    }
    .segmented button:last-child { border-right: 0; }
    .segmented button.active {
      background: #e8f0fe;
      color: #174ea6;
    }
    canvas {
      display: block;
      width: 100%;
      height: 420px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: white;
    }
    .chart-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
      align-items: start;
    }
    .chart-card {
      min-width: 0;
    }
    .chart-card-head {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 8px;
      margin-bottom: 8px;
    }
    .chart-card h3 {
      margin: 0;
      font-size: 14px;
      letter-spacing: 0;
    }
    .stack {
      display: grid;
      gap: 12px;
    }
    .grid2 {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 8px;
    }
    .grid3 {
      display: grid;
      grid-template-columns: 1fr 1fr 1fr;
      gap: 8px;
    }
    .spot-entry-grid {
      display: grid;
      grid-template-columns: repeat(5, minmax(78px, 1fr));
      gap: 8px;
      min-width: 420px;
    }
    .spot-entry-groups {
      display: grid;
      gap: 14px;
      margin-top: 12px;
      overflow-x: auto;
      padding-bottom: 2px;
    }
    .spot-entry-section {
      display: grid;
      gap: 8px;
    }
    .spot-entry-title {
      color: var(--ink);
      font-size: 13px;
      font-weight: 700;
      line-height: 1.4;
    }
    .hint, .status {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.5;
    }
    .status strong { color: var(--ink); }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
    }
    th, td {
      border-bottom: 1px solid var(--line);
      padding: 7px 6px;
      text-align: left;
      vertical-align: top;
    }
    th {
      color: var(--muted);
      font-weight: 600;
      background: #fafafa;
      position: sticky;
      top: 0;
    }
    .table-wrap {
      max-height: 260px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 6px;
    }
    .overview-wrap {
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 6px;
      margin: 8px 0 12px;
      background: #fff;
    }
    .overview-table {
      min-width: 2400px;
      white-space: nowrap;
    }
    .overview-table th, .overview-table td {
      text-align: right;
      padding: 6px 8px;
    }
    .overview-table th:first-child, .overview-table td:first-child {
      text-align: left;
      position: sticky;
      left: 0;
      background: #fff;
      z-index: 1;
    }
    .overview-table thead th:first-child {
      background: #fafafa;
      z-index: 2;
    }
    .overview-table .group-head {
      text-align: center;
      border-left: 1px solid var(--line);
    }
    .overview-table .muted-cell {
      color: var(--muted);
    }
    #tooltip {
      position: fixed;
      display: none;
      pointer-events: none;
      background: rgba(32,33,36,.94);
      color: #fff;
      border-radius: 6px;
      padding: 8px 10px;
      font-size: 12px;
      line-height: 1.45;
      z-index: 20;
      max-width: 260px;
    }
    @media (max-width: 980px) {
      .layout, .topbar {
        grid-template-columns: 1fr;
      }
      .chart-grid {
        grid-template-columns: repeat(2, minmax(0, 1fr));
      }
      canvas { height: 420px; }
    }
  </style>
</head>
<body>
  <header>
    <h1>钢材 05 / 10 / 01 合约基差季节性</h1>
  </header>

  <main>
    <section class="topbar">
      <div class="metric"><div class="label">现货最新日期</div><div class="value" id="mLatestDate">-</div></div>
      <div class="metric"><div class="label">上海热卷价格</div><div class="value" id="mSpot">-</div></div>
      <div class="metric"><div class="label">现货历史条数</div><div class="value" id="mSpotCount">-</div></div>
      <div class="metric"><div class="label">期货收盘条数</div><div class="value" id="mFutCount">-</div></div>
    </section>

    <section class="layout">
      <div class="panel">
        <div class="chart-toolbar">
          <div>
            <h2>基差季节性图</h2>
            <div class="hint">普方坯、广州韶钢、中心库、杭州中天、乐从热卷、上海热卷；基差 = 现货价 - 对应合约收盘价</div>
          </div>
          <div class="row">
            <div class="segmented" id="monthTabs">
              <button data-month="5">05</button>
              <button data-month="10" class="active">10</button>
              <button data-month="1">01</button>
            </div>
          </div>
        </div>
        <div class="overview-wrap">
          <table class="overview-table">
            <thead id="overviewHead"></thead>
            <tbody id="overviewBody">
              <tr><td colspan="13" class="muted-cell">加载中</td></tr>
            </tbody>
          </table>
        </div>
        <div class="chart-grid">
          <section class="chart-card">
            <div class="chart-card-head">
              <div>
                <h3>普方坯基差季节性</h3>
                <div class="hint" id="billetChartHint">基差 = 普方坯现货价 - 对应 RB 合约收盘价</div>
              </div>
            </div>
            <canvas id="billetBasisChart" width="560" height="420"></canvas>
            <div class="status" id="billetChartStatus"></div>
          </section>
          <section class="chart-card">
            <div class="chart-card-head">
              <div>
                <h3>广州韶钢基差季节性</h3>
                <div class="hint" id="guangzhouShaogangChartHint">基差 = 广州韶钢现货价 - 对应 RB 合约收盘价</div>
              </div>
            </div>
            <canvas id="guangzhouShaogangBasisChart" width="560" height="420"></canvas>
            <div class="status" id="guangzhouShaogangChartStatus"></div>
          </section>
          <section class="chart-card">
            <div class="chart-card-head">
              <div>
                <h3>中心库基差季节性</h3>
                <div class="hint" id="centerWarehouseChartHint">基差 = 中心库现货价 - 对应 RB 合约收盘价</div>
              </div>
            </div>
            <canvas id="centerWarehouseBasisChart" width="560" height="420"></canvas>
            <div class="status" id="centerWarehouseChartStatus"></div>
          </section>
          <section class="chart-card">
            <div class="chart-card-head">
              <div>
                <h3>杭州中天基差季节性</h3>
                <div class="hint" id="hangzhouZhongtianChartHint">基差 = 杭州中天现货价 - 对应 RB 合约收盘价</div>
              </div>
            </div>
            <canvas id="hangzhouZhongtianBasisChart" width="560" height="420"></canvas>
            <div class="status" id="hangzhouZhongtianChartStatus"></div>
          </section>
          <section class="chart-card">
            <div class="chart-card-head">
              <div>
                <h3>乐从热卷基差季节性</h3>
                <div class="hint" id="lecongHotcoilChartHint">基差 = 乐从热卷现货价 - 对应 HC 合约收盘价</div>
              </div>
            </div>
            <canvas id="lecongHotcoilBasisChart" width="560" height="420"></canvas>
            <div class="status" id="lecongHotcoilChartStatus"></div>
          </section>
          <section class="chart-card">
            <div class="chart-card-head">
              <div>
                <h3>上海热卷基差季节性</h3>
                <div class="hint" id="shanghaiHotcoilChartHint">基差 = 上海热卷现货价 - 对应 HC 合约收盘价</div>
              </div>
            </div>
            <canvas id="shanghaiHotcoilBasisChart" width="560" height="420"></canvas>
            <div class="status" id="shanghaiHotcoilChartStatus"></div>
          </section>
        </div>
      </div>

      <aside class="stack">
        <div class="panel">
          <h2>新增 / 更新日度数据</h2>
          <div class="grid2">
            <div>
              <label for="dateInput">日期</label>
              <input id="dateInput" type="date">
            </div>
          </div>
          <div class="spot-entry-groups" id="spotEntryGrid"></div>
          <div class="row" style="margin-top: 8px;">
            <button class="primary" id="saveSpotBtn">保存全部现货</button>
            <button id="latestTradeBtn">识别最新交易日</button>
          </div>

          <div class="grid3" style="margin-top: 12px;">
            <div>
              <label for="fMonthInput">合约月</label>
              <select id="fMonthInput">
                <option value="5">05</option>
                <option value="10" selected>10</option>
                <option value="1">01</option>
              </select>
            </div>
            <div>
              <label for="contractInput">合约代码</label>
              <input id="contractInput" placeholder="HC2510">
            </div>
            <div>
              <label for="futuresInput">期货收盘</label>
              <input id="futuresInput" type="number" step="1">
            </div>
          </div>
          <div class="row" style="margin-top: 8px;">
            <button id="fetchFuturesBtn">自动获取收盘</button>
            <button class="primary" id="saveFuturesBtn">保存期货</button>
          </div>
          <div class="status" id="formStatus" style="margin-top: 8px;"></div>
        </div>

        <div class="panel">
          <h2>批量补近期合约收盘</h2>
          <div class="hint" style="margin-bottom: 8px;">自动补数优先使用新浪期货日 K，失败时回退上期所日行情。</div>
          <div class="grid2">
            <div>
              <label for="batchMonth">合约月</label>
              <select id="batchMonth">
                <option value="5">05</option>
                <option value="10" selected>10</option>
                <option value="1">01</option>
              </select>
            </div>
            <div>
              <label for="batchLimit">最多日期数</label>
              <input id="batchLimit" type="number" value="20" min="1" max="250">
            </div>
          </div>
          <div class="row" style="margin-top: 8px;">
            <button id="fetchMissingBtn">补当前合约月</button>
            <button class="primary" id="fetchMissingAllBtn">补 05/10/01</button>
            <button id="importExcelBtn">从 Excel 同步</button>
          </div>
          <div class="hint" id="batchStatus" style="margin-top: 8px;"></div>
        </div>

        <div class="panel">
          <h2>最近数据</h2>
          <div class="table-wrap">
            <table>
              <thead><tr><th>日期</th><th>现货</th><th>期货</th></tr></thead>
              <tbody id="recentBody"></tbody>
            </table>
          </div>
        </div>
      </aside>
    </section>
  </main>
  <div id="tooltip"></div>

<script>
const state = {
  month: 10,
  activeChart: 'centerWarehouse'
};

const SPOT_CONFIG = {
  shanghai_hotcoil: {label: '上海热卷', product: 'HC'},
  lecong_hotcoil: {label: '乐从热卷', product: 'HC'},
  tangshan_hotcoil: {label: '唐山热卷', product: 'HC'},
  zhongtian_factory: {label: '中天厂发', product: 'RB'},
  center_warehouse: {label: '中心库', product: 'RB'},
  guangzhou_shaogang: {label: '广州韶钢', product: 'RB'},
  billet: {label: '普方坯', product: 'RB'},
  hangzhou_zhongtian: {label: '杭州中天', product: 'RB'}
};

const SPOT_KEYS = Object.keys(SPOT_CONFIG);
const OVERVIEW_MONTHS = ['05', '10', '01'];
const SPOT_ENTRY_GROUPS = [
  {
    title: '长材组',
    keys: ['billet', 'hangzhou_zhongtian', 'zhongtian_factory', 'center_warehouse', 'guangzhou_shaogang']
  },
  {
    title: '板材组',
    keys: ['lecong_hotcoil', 'shanghai_hotcoil', 'tangshan_hotcoil']
  }
];
const CHARTS = {
  billet: {
    title: '普方坯',
    spotKey: 'billet',
    product: 'RB',
    canvasId: 'billetBasisChart',
    hintId: 'billetChartHint',
    statusId: 'billetChartStatus',
    points: [],
    hoverPoints: []
  },
  guangzhouShaogang: {
    title: '广州韶钢',
    spotKey: 'guangzhou_shaogang',
    product: 'RB',
    canvasId: 'guangzhouShaogangBasisChart',
    hintId: 'guangzhouShaogangChartHint',
    statusId: 'guangzhouShaogangChartStatus',
    points: [],
    hoverPoints: []
  },
  centerWarehouse: {
    title: '中心库',
    spotKey: 'center_warehouse',
    product: 'RB',
    canvasId: 'centerWarehouseBasisChart',
    hintId: 'centerWarehouseChartHint',
    statusId: 'centerWarehouseChartStatus',
    points: [],
    hoverPoints: []
  },
  hangzhouZhongtian: {
    title: '杭州中天',
    spotKey: 'hangzhou_zhongtian',
    product: 'RB',
    canvasId: 'hangzhouZhongtianBasisChart',
    hintId: 'hangzhouZhongtianChartHint',
    statusId: 'hangzhouZhongtianChartStatus',
    points: [],
    hoverPoints: []
  },
  lecongHotcoil: {
    title: '乐从热卷',
    spotKey: 'lecong_hotcoil',
    product: 'HC',
    canvasId: 'lecongHotcoilBasisChart',
    hintId: 'lecongHotcoilChartHint',
    statusId: 'lecongHotcoilChartStatus',
    points: [],
    hoverPoints: []
  },
  shanghaiHotcoil: {
    title: '上海热卷',
    spotKey: 'shanghai_hotcoil',
    product: 'HC',
    canvasId: 'shanghaiHotcoilBasisChart',
    hintId: 'shanghaiHotcoilChartHint',
    statusId: 'shanghaiHotcoilChartStatus',
    points: [],
    hoverPoints: []
  }
};

const $ = (id) => document.getElementById(id);

async function api(path, options = {}) {
  const response = await fetch(path, {
    headers: {'Content-Type': 'application/json'},
    ...options
  });
  const data = await response.json();
  if (!response.ok || data.error) throw new Error(data.error || response.statusText);
  return data;
}

function fmtNumber(value) {
  if (value === null || value === undefined || Number.isNaN(Number(value))) return '-';
  return Number(value).toLocaleString('zh-CN', {maximumFractionDigits: 0});
}

function productPrefix() {
  return CHARTS[state.activeChart].product;
}

function spotLabel() {
  return SPOT_CONFIG[CHARTS[state.activeChart].spotKey].label;
}

function spotInputId(spotKey) {
  return `spotInput_${spotKey}`;
}

function buildSpotControls() {
  $('spotEntryGrid').innerHTML = SPOT_ENTRY_GROUPS.map(group => (
    `<section class="spot-entry-section" aria-label="${group.title}">
      <div class="spot-entry-title">${group.title}</div>
      <div class="spot-entry-grid">
        ${group.keys.map(key => (
          `<div><label for="${spotInputId(key)}">${SPOT_CONFIG[key].label}现货</label><input id="${spotInputId(key)}" type="number" step="1"></div>`
        )).join('')}
      </div>
    </section>`
  )).join('');

  const groupHeaders = SPOT_KEYS.map(key => (
    `<th class="group-head" colspan="4">${SPOT_CONFIG[key].label}</th>`
  )).join('');
  const subHeaders = SPOT_KEYS.map(() => (
    '<th>现货</th><th>05</th><th>10</th><th>01</th>'
  )).join('');
  $('overviewHead').innerHTML = `<tr><th rowspan="2">日期</th>${groupHeaders}</tr><tr>${subHeaders}</tr>`;
}

function nextContractCode(dateText, product, month) {
  const d = new Date(dateText + 'T00:00:00');
  let year = d.getFullYear();
  if ((d.getMonth() + 1) > Number(month)) year += 1;
  return `${product}${String(year).slice(2)}${String(month).padStart(2, '0')}`;
}

function updateContractInput() {
  const dateText = $('dateInput').value;
  const month = Number($('fMonthInput').value);
  if (!dateText) return;
  $('contractInput').value = nextContractCode(dateText, productPrefix(), month);
}

async function refreshState() {
  const active = CHARTS[state.activeChart];
  const data = await api(`/api/state?spot_key=${encodeURIComponent(active.spotKey)}`);
  active.product = data.product || productPrefix();
  $('mLatestDate').textContent = data.latest_spot ? data.latest_spot.date : '-';
  $('mSpot').textContent = data.latest_spot ? fmtNumber(data.latest_spot.price) : '-';
  $('mSpotCount').textContent = fmtNumber(data.spot_count);
  $('mFutCount').textContent = fmtNumber(data.futures_count);
  if (!$('dateInput').value) $('dateInput').value = data.default_date;
  await refreshSpotPlaceholders();
  updateContractInput();
  await refreshOverview();
  await refreshRecent();
}

async function refreshSpotPlaceholders() {
  for (const spotKey of SPOT_KEYS) {
    const data = await api(`/api/state?spot_key=${encodeURIComponent(spotKey)}`);
    if (data.latest_spot) {
      $(spotInputId(spotKey)).placeholder = String(data.latest_spot.price);
    }
  }
}

function fmtSigned(value) {
  if (value === null || value === undefined || Number.isNaN(Number(value))) return '-';
  const num = Number(value);
  return `${num > 0 ? '+' : ''}${num.toLocaleString('zh-CN', {maximumFractionDigits: 0})}`;
}

async function refreshOverview() {
  const data = await api('/api/recent-overview?limit=5');
  const colspan = 1 + SPOT_KEYS.length * 4;
  if (!data.rows.length) {
    $('overviewBody').innerHTML = `<tr><td colspan="${colspan}" class="muted-cell">暂无最近五日共同报价</td></tr>`;
    return;
  }
  $('overviewBody').innerHTML = data.rows.map(row => {
    const cells = [`<td>${row.date}</td>`];
    for (const key of SPOT_KEYS) {
      const series = row.series[key] || {};
      cells.push(`<td>${fmtNumber(series.price)}</td>`);
      for (const month of OVERVIEW_MONTHS) {
        const item = (series.basis || {})[month] || {};
        cells.push(`<td title="${item.contract_code || ''}">${fmtSigned(item.basis)}</td>`);
      }
    }
    return `<tr>${cells.join('')}</tr>`;
  }).join('');
}

async function refreshRecent() {
  const active = CHARTS[state.activeChart];
  const data = await api(`/api/recent?limit=12&spot_key=${encodeURIComponent(active.spotKey)}`);
  $('recentBody').innerHTML = data.rows.map(row => (
    `<tr><td>${row.date}</td><td>${fmtNumber(row.price)}</td><td>${row.futures || ''}</td></tr>`
  )).join('');
}

async function refreshChart(chartKey) {
  const chart = CHARTS[chartKey];
  const data = await api(`/api/basis?product=${chart.product}&contract_month=${state.month}&spot_key=${encodeURIComponent(chart.spotKey)}`);
  chart.points = data.points;
  const month = String(state.month).padStart(2, '0');
  $(chart.hintId).textContent = `基差 = ${SPOT_CONFIG[chart.spotKey].label}现货价 - ${chart.product}${month} 对应合约收盘价；横轴为 ${month}-15 至次年 ${month}-15`;
  drawChart(chartKey);
}

async function refreshCharts() {
  await Promise.all(Object.keys(CHARTS).map(refreshChart));
}

function groupByYear(points) {
  const groups = new Map();
  for (const point of points) {
    const key = point.season_label || String(point.year);
    if (!groups.has(key)) groups.set(key, []);
    groups.get(key).push(point);
  }
  return Array.from(groups.entries()).sort((a, b) => String(a[0]).localeCompare(String(b[0])));
}

function formatAxisDate(startIso, offsetDays) {
  const d = new Date(startIso + 'T00:00:00');
  d.setDate(d.getDate() + offsetDays);
  return `${String(d.getMonth() + 1).padStart(2, '0')}-${String(d.getDate()).padStart(2, '0')}`;
}

function drawChart(chartKey) {
  const chart = CHARTS[chartKey];
  const canvas = $(chart.canvasId);
  const ctx = canvas.getContext('2d');
  const rect = canvas.getBoundingClientRect();
  const ratio = window.devicePixelRatio || 1;
  canvas.width = Math.round(rect.width * ratio);
  canvas.height = Math.round(rect.height * ratio);
  ctx.setTransform(ratio, 0, 0, ratio, 0, 0);
  const w = rect.width;
  const h = rect.height;
  ctx.clearRect(0, 0, w, h);

  const margin = {left: 58, right: 18, top: 58, bottom: 46};
  const plotW = w - margin.left - margin.right;
  const plotH = h - margin.top - margin.bottom;
  const points = chart.points;
  chart.hoverPoints = [];

  ctx.fillStyle = '#fff';
  ctx.fillRect(0, 0, w, h);

  if (!points.length) {
    ctx.fillStyle = '#667085';
    ctx.font = '14px sans-serif';
    ctx.fillText('还没有可绘制的基差数据。请先录入或自动获取对应合约收盘价。', margin.left, margin.top + 24);
    $(chart.statusId).innerHTML = `当前图表需要同一日期同时存在${SPOT_CONFIG[chart.spotKey].label}现货价和期货收盘价。`;
    return;
  }

  const minY = Math.min(...points.map(p => p.basis));
  const maxY = Math.max(...points.map(p => p.basis));
  const pad = Math.max(20, (maxY - minY) * 0.12);
  const yMin = Math.floor((minY - pad) / 50) * 50;
  const yMax = Math.ceil((maxY + pad) / 50) * 50;
  const maxWindowDay = Math.max(...points.map(p => p.window_total_days || 365));
  const axisStart = points[0].window_start;
  const x = day => margin.left + day / (maxWindowDay || 365) * plotW;
  const y = val => margin.top + (yMax - val) / (yMax - yMin || 1) * plotH;

  ctx.strokeStyle = '#d9dee7';
  ctx.lineWidth = 1;
  ctx.fillStyle = '#667085';
  ctx.font = '12px sans-serif';

  const yTicks = 6;
  for (let i = 0; i <= yTicks; i++) {
    const val = yMin + (yMax - yMin) * i / yTicks;
    const yy = y(val);
    ctx.beginPath();
    ctx.moveTo(margin.left, yy);
    ctx.lineTo(w - margin.right, yy);
    ctx.stroke();
    ctx.fillText(String(Math.round(val)), 8, yy + 4);
  }

  const xTicks = [0, 61, 122, 183, 244, 305, maxWindowDay].filter((day, index, arr) => day <= maxWindowDay && arr.indexOf(day) === index);
  for (const day of xTicks) {
    const label = formatAxisDate(axisStart, day);
    const xx = x(day);
    ctx.beginPath();
    ctx.moveTo(xx, margin.top);
    ctx.lineTo(xx, h - margin.bottom);
    ctx.stroke();
    ctx.fillText(label, xx - 15, h - 18);
  }

  ctx.strokeStyle = '#202124';
  ctx.beginPath();
  ctx.moveTo(margin.left, margin.top);
  ctx.lineTo(margin.left, h - margin.bottom);
  ctx.lineTo(w - margin.right, h - margin.bottom);
  ctx.stroke();

  const colors = ['#1f6feb', '#248a3d', '#c2410c', '#7c3aed', '#b7791f', '#0891b2', '#be185d', '#475569'];
  const groups = groupByYear(points);
  groups.forEach(([label, rows], idx) => {
    rows.sort((a, b) => a.window_day - b.window_day);
    const color = colors[idx % colors.length];
    ctx.strokeStyle = color;
    ctx.lineWidth = idx === groups.length - 1 ? 2.6 : 1.7;
    ctx.beginPath();
    rows.forEach((p, i) => {
      const xx = x(p.window_day);
      const yy = y(p.basis);
      if (i === 0) ctx.moveTo(xx, yy);
      else ctx.lineTo(xx, yy);
      chart.hoverPoints.push({...p, x: xx, y: yy, color, chartKey});
    });
    ctx.stroke();

  });

  let lx = margin.left;
  let ly = 18;
  ctx.font = '12px sans-serif';
  groups.forEach(([seasonLabel], idx) => {
    const label = String(seasonLabel);
    const itemWidth = Math.ceil(ctx.measureText(label).width) + 30;
    if (lx + itemWidth > w - margin.right) {
      lx = margin.left;
      ly += 18;
    }
    ctx.fillStyle = colors[idx % colors.length];
    ctx.fillRect(lx, ly - 8, 10, 10);
    ctx.fillStyle = '#202124';
    ctx.fillText(label, lx + 14, ly + 1);
    lx += itemWidth;
  });

  const latest = points[points.length - 1];
  $(chart.statusId).innerHTML = `已绘制 <strong>${points.length}</strong> 个基差点；最新：<strong>${latest.date}</strong>，${latest.contract_code} 基差 <strong>${fmtNumber(latest.basis)}</strong>。`;
}

function attachEvents() {
  $('monthTabs').addEventListener('click', async (event) => {
    if (!event.target.dataset.month) return;
    state.month = Number(event.target.dataset.month);
    document.querySelectorAll('#monthTabs button').forEach(btn => btn.classList.toggle('active', btn === event.target));
    await refreshCharts();
  });
  $('dateInput').addEventListener('change', updateContractInput);
  $('fMonthInput').addEventListener('change', updateContractInput);

  $('saveSpotBtn').addEventListener('click', () => withBusy('saveSpotBtn', async () => {
    const prices = {};
    for (const spotKey of SPOT_KEYS) {
      prices[spotKey] = $(spotInputId(spotKey)).value;
    }
    const data = await api('/api/spots', {method: 'POST', body: JSON.stringify({date: $('dateInput').value, prices})});
    const savedLabels = Object.keys(data.saved).map(key => SPOT_CONFIG[key].label).join('、');
    $('formStatus').textContent = `${savedLabels}现货已保存。`;
    for (const spotKey of SPOT_KEYS) {
      $(spotInputId(spotKey)).value = '';
    }
    await refreshState();
    await refreshCharts();
  }));

  $('saveFuturesBtn').addEventListener('click', () => withBusy('saveFuturesBtn', async () => {
    await api('/api/futures', {method: 'POST', body: JSON.stringify({
      date: $('dateInput').value,
      product: productPrefix(),
      contract_month: Number($('fMonthInput').value),
      contract_code: $('contractInput').value,
      close: $('futuresInput').value
    })});
    $('formStatus').textContent = '期货收盘已保存。';
    await refreshState();
    await refreshCharts();
  }));

  $('fetchFuturesBtn').addEventListener('click', () => withBusy('fetchFuturesBtn', async () => {
    const data = await api('/api/fetch-futures', {method: 'POST', body: JSON.stringify({
      date: $('dateInput').value,
      product: productPrefix(),
      contract_month: Number($('fMonthInput').value),
      contract_code: $('contractInput').value
    })});
    $('futuresInput').value = data.close;
    $('formStatus').textContent = `${data.contract_code} 收盘价已获取并保存：${data.close}（${data.source}）`;
    await refreshState();
    await refreshCharts();
  }));

  $('latestTradeBtn').addEventListener('click', () => withBusy('latestTradeBtn', async () => {
    const data = await api('/api/latest-trading-date');
    $('dateInput').value = data.date;
    updateContractInput();
    $('formStatus').textContent = `最新可识别交易日：${data.date}`;
  }));

  $('fetchMissingBtn').addEventListener('click', () => withBusy('fetchMissingBtn', async () => {
    const data = await api('/api/fetch-missing', {method: 'POST', body: JSON.stringify({
      product: productPrefix(),
      spot_key: CHARTS[state.activeChart].spotKey,
      contract_month: Number($('batchMonth').value),
      limit: Number($('batchLimit').value)
    })});
    $('batchStatus').textContent = `尝试 ${data.attempted} 天，保存 ${data.saved} 条，失败 ${data.error_count || 0} 条。${data.errors.length ? '示例错误：' + data.errors.join('；') : ''}`;
    await refreshState();
    await refreshCharts();
  }));

  $('fetchMissingAllBtn').addEventListener('click', () => withBusy('fetchMissingAllBtn', async () => {
    const data = await api('/api/fetch-missing-all', {method: 'POST', body: JSON.stringify({
      product: productPrefix(),
      spot_key: CHARTS[state.activeChart].spotKey,
      limit: Number($('batchLimit').value)
    })});
    $('batchStatus').textContent = `尝试 ${data.attempted_dates} 天，保存 ${data.saved} 条，跳过已有 ${data.skipped_existing} 条，失败 ${data.error_count || 0} 条。${data.errors.length ? '示例错误：' + data.errors.join('；') : ''}`;
    await refreshState();
    await refreshCharts();
  }));

  $('importExcelBtn').addEventListener('click', () => withBusy('importExcelBtn', async () => {
    const data = await api('/api/import-excel', {method: 'POST', body: '{}'});
    $('batchStatus').textContent = `Excel 已同步：现货 ${data.spot.imported} 条，期货 ${data.futures.imported} 条，跳过 ${data.skipped} 行。`;
    await refreshState();
    await refreshCharts();
  }));

  for (const [chartKey, chart] of Object.entries(CHARTS)) {
    const canvas = $(chart.canvasId);
    canvas.addEventListener('mousemove', event => showTooltip(event, chartKey));
    canvas.addEventListener('mouseleave', () => $('tooltip').style.display = 'none');
    canvas.addEventListener('mouseenter', () => {
      state.activeChart = chartKey;
      updateContractInput();
    });
  }
  window.addEventListener('resize', () => Object.keys(CHARTS).forEach(drawChart));
}

async function withBusy(buttonId, fn) {
  const button = $(buttonId);
  button.disabled = true;
  try {
    await fn();
  } catch (error) {
    $('formStatus').textContent = error.message;
    $('batchStatus').textContent = error.message;
  } finally {
    button.disabled = false;
  }
}

function showTooltip(event, chartKey) {
  const chart = CHARTS[chartKey];
  if (!chart.hoverPoints.length) return;
  const rect = $(chart.canvasId).getBoundingClientRect();
  const mx = event.clientX - rect.left;
  const my = event.clientY - rect.top;
  let best = null;
  let bestDist = Infinity;
  for (const p of chart.hoverPoints) {
    const dist = Math.hypot(p.x - mx, p.y - my);
    if (dist < bestDist) {
      bestDist = dist;
      best = p;
    }
  }
  const tip = $('tooltip');
  if (!best || bestDist > 18) {
    tip.style.display = 'none';
    return;
  }
  tip.innerHTML = `${best.date} ${best.contract_code}<br>基差：${fmtNumber(best.basis)}<br>现货：${fmtNumber(best.spot)}，期货：${fmtNumber(best.futures_close)}`;
  tip.style.left = `${event.clientX + 12}px`;
  tip.style.top = `${event.clientY + 12}px`;
  tip.style.display = 'block';
}

(async function init() {
  buildSpotControls();
  attachEvents();
  await refreshState();
  await refreshCharts();
})();
</script>
</body>
</html>
"""


def bootstrap() -> None:
    init_db()
    if EXCEL_PATH.exists():
        result = sync_from_excel(EXCEL_PATH)
        print(f"Synced business data from Excel: {result}", flush=True)


def main() -> None:
    bootstrap()
    server = ThreadingHTTPServer((DEFAULT_HOST, DEFAULT_PORT), AppHandler)
    print(f"Dashboard: http://{DEFAULT_HOST}:{DEFAULT_PORT}", flush=True)
    print(f"Excel: {EXCEL_PATH}", flush=True)
    print(f"Database: {DB_PATH}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    if "--sync-sqlite-to-excel" in sys.argv:
        print(json.dumps(sync_sqlite_business_data_to_excel(), ensure_ascii=False, indent=2))
    else:
        main()
